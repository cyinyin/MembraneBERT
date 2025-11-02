'''
fun: Read article links from Excel
'''

import os
from unit.path import Args
import openpyxl

doi_path = Args().MOF_COF
doi_xlsx = openpyxl.load_workbook(doi_path, read_only=True)
sheets = doi_xlsx.sheetnames
sht = sheets[0]
sheet_data = doi_xlsx[sht]
paper_doi=[]


def get_doi():
    for all_doi in sheet_data.iter_rows(min_row=2, min_col=33, max_col=33):
        for doi in all_doi:
            paper_doi.append(doi.value)
    return paper_doi

# if __name__ == '__main__':
#     print(get_doi())
