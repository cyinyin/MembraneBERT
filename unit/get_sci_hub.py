# -*- coding: utf-8 -*-
import time
from selenium.webdriver.chrome.service import Service
from unit.read_excel_dio import get_doi

from selenium import webdriver
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
import selenium.common.exceptions as Exceptions
from selenium.common.exceptions import WebDriverException
from selenium.webdriver.chrome.options import Options
import fitz
import os
import re
import pandas as pd
from openpyxl import load_workbook

import datetime
import subprocess


def init_driver(address):
    chrome_options = Options()
    # chrome_options.add_argument('--headless')
    chrome_options.add_argument('--incognito')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument("--test-type")
    chrome_options.add_argument("--disable-popup-blocking")
    # chrome_options.add_argument('blink-settings=imagesEnabled=false')
    chrome_options.add_experimental_option('excludeSwitches',
                                           ['enable-automation'])
    chrome_options.page_load_strategy = 'eager'
    chrome_options.add_experimental_option("useAutomationExtension", False)
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])

    download_dir = "D://ppdf"
    chrome_options.add_experimental_option('prefs', {
        'download.default_directory': download_dir,
        'download.prompt_for_download': False,
        'download.directory_upgrade': True,
        'plugins.always_open_pdf_externally': True
    })

    try:
        chrome_options = Options()
        service = Service(executable_path=address)
        driver = webdriver.Chrome(service=service, options=chrome_options)

    except Exceptions.SessionNotCreatedException as e:
        print(f"error: {e}")
        raise

    return driver


def build_graph_from_title():
    download_dir = "D://ppdf"
    desired_capabilities = DesiredCapabilities.CHROME
    desired_capabilities["pageLoadStrategy"] = "none"
    driver_path = 'D:/chromedriver-win64/chromedriver.exe'
    for url in get_doi():
        driver = init_driver(driver_path)
        sentences_with_values = []
        file_name = url.split('/')[-1]
        new_file=f"{file_name}.pdf"
        driver.get('https://sci-hub.shop/')
        search_bar = WebDriverWait(driver=driver, timeout=30, poll_frequency=0.5).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="input"]/form/input[2]'))
        )
        search_bar.clear()
        search_bar.send_keys(url)

        WebDriverWait(driver=driver, timeout=30, poll_frequency=0.5).until(
            EC.visibility_of_element_located(
                (By.XPATH,  '//*[@id="open"]')
            )
        ).click()

        WebDriverWait(driver=driver, timeout=60, poll_frequency=0.5).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="buttons"]/button[2] '))
        ).click()
        time.sleep(60)
        driver.close()
        files = os.listdir(download_dir)
        tmp_files = [f for f in files if f.endswith('.tmp')]

        if not tmp_files:
            print("No .tmp files found in the directory.")
        else:
            tmp_file_paths = [os.path.join(download_dir, f) for f in tmp_files]

            latest_tmp = max(tmp_file_paths, key=os.path.getmtime)

            new_file_path = os.path.join(download_dir, new_file)
            os.rename(latest_tmp, new_file_path)
            print(f"The latest .tmp file has been renamed to: {new_file_path}")

            pdf_document = fitz.open(new_file_path)

            keyword = 'diffusivity'
            found = False
            output_excel = "D:/MembraneBERT/data/paragraph_value.xlsx"
            value_pattern = re.compile(rf"{keyword}.*?([-+]?\d*\.?\d*)\s*×?\s*10\^([-+]?\d+)", re.IGNORECASE)

            for page_num in range(pdf_document.page_count):
                page = pdf_document.load_page(page_num)
                text = page.get_text()

                sentences = re.split(r'(?<!\w\.\w.)(?<![A-Z][a-z]\.)(?<=\.|\?)\s', text)
                for sentence in sentences:
                    if keyword.lower() in sentence.lower():
                        found=True
                        match = value_pattern.search(sentence)
                        if match:
                            base_value = float(match.group(1))
                            exponent = int(match.group(2))
                            scientific_value = base_value * (10 ** exponent)

                            sentences_with_values.append((url, sentence.strip(), scientific_value))
                        else:
                            scientific_value=None
                            sentences_with_values.append((url, sentence.strip(), scientific_value))
                            print(f"Sentence: {sentence}\n not found Value\n")
            if sentences_with_values:
                df = pd.DataFrame(sentences_with_values, columns=["Paper url", "Sentence", "Calculated Value"])
                try:
                    book = load_workbook(output_excel)
                    writer = pd.ExcelWriter(output_excel, engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    start_row = writer.sheets['Sheet1'].max_row
                    df.to_excel(writer, startrow=start_row, header=False, index=False)
                    writer.save()
                except FileNotFoundError:
                    df.to_excel(output_excel, index=False)
                    print(f"Created new Excel file: {output_excel}")
            if not found:
                print(f"Keyword '{keyword}' not found in the PDF.")


if __name__ == '__main__':
    build_graph_from_title()
    pass